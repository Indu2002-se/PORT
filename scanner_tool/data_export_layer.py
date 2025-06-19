"""
Data Export Layer - Handles exporting scan results to different formats
"""

import os
import logging
import csv
import json
import re
from typing import Dict, List, Optional, Any, cast
from datetime import datetime
import uuid
from supabase import create_client, Client
from dotenv import load_dotenv

# Import third-party libraries for specific file formats
try:
    import openpyxl
    import openpyxl.styles
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    EXCEL_AVAILABLE = True
    # Type hints for proper type checking
    from typing import TYPE_CHECKING
    if TYPE_CHECKING:
        from openpyxl.workbook.workbook import Workbook as WorkbookType
        from openpyxl.worksheet.worksheet import Worksheet as WorksheetType
except ImportError:
    openpyxl = None
    Workbook = None
    Worksheet = None
    EXCEL_AVAILABLE = False

try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
except ImportError:
    FPDF = None
    PDF_AVAILABLE = False

from colorama import Fore

# Load environment variables
load_dotenv()

logger = logging.getLogger(__name__)

# Initialize Supabase client with hardcoded credentials (for development only)
supabase_url = "https://tuhtempenltbwzjhrzmx.supabase.co"
supabase_key = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InR1aHRlbXBlbmx0Ynd6amhyem14Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NTAzMjE5NjgsImV4cCI6MjA2NTg5Nzk2OH0.9eqF00LfCu1k_u8zxVEKhQs3bYwgVbvQ5pzW1zuzvAw"

def get_supabase() -> Client:
    """Get Supabase client instance."""
    return create_client(supabase_url, supabase_key)

class DataExportLayer:
    """
    Handles exporting scan results to various formats (Excel, CSV, PDF, JSON).
    """

    def __init__(self):
        """Initialize the data export layer."""
        self.export_dir = "scan_results"

        # Create export directory if it doesn't exist
        if not os.path.exists(self.export_dir):
            try:
                os.makedirs(self.export_dir)
            except OSError as e:
                logger.error(f"Error creating export directory: {e}")
                # Fallback to current directory
                self.export_dir = ""

    def _prepare_data(self, scan_results: Dict[int, Dict], host: str) -> List[List[str]]:
        """
        Prepare scan data for export.

        Args:
            scan_results: Dictionary of open ports and their data including service and banner info
            host: The hostname or IP address scanned

        Returns:
            List[List[str]]: List of rows for export
        """
        # Create header row
        data = [["Host", "Port", "Status", "Service", "Version", "Server", "Banner", "SSL Certificate", "Scan Date"]]

        # Add scan timestamp
        scan_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Add data rows
        for port, port_data in scan_results.items():
            service = port_data.get("service", "")
            version = port_data.get("version", "")
            server = port_data.get("server", "")
            banner = port_data.get("banner", "")

            # Clean and truncate long banners
            banner = banner.replace('\r\n', ' | ').replace('\n', ' | ').replace('\r', ' | ')
            if len(banner) > 500:
                banner = banner[:497] + "..."

            # Format SSL certificate information
            ssl_cert = port_data.get("ssl_cert", {})
            ssl_info = ""
            if ssl_cert and isinstance(ssl_cert, dict) and any(ssl_cert.values()):
                cert_details = []
                if ssl_cert.get("issued_to"):
                    cert_details.append(f"Issued To: {ssl_cert['issued_to']}")
                if ssl_cert.get("issued_by"):
                    cert_details.append(f"Issued By: {ssl_cert['issued_by']}")
                if ssl_cert.get("valid_from"):
                    cert_details.append(f"Valid From: {ssl_cert['valid_from']}")
                if ssl_cert.get("valid_until"):
                    cert_details.append(f"Valid Until: {ssl_cert['valid_until']}")
                if ssl_cert.get("version"):
                    cert_details.append(f"Version: {ssl_cert['version']}")

                ssl_info = ", ".join(cert_details)

            data.append([
                host, 
                str(port), 
                "Open", 
                service, 
                version, 
                server, 
                banner, 
                ssl_info, 
                scan_time
            ])

        return data

    def export_to_csv(self, scan_results: Dict[int, Dict], host: str, filename: Optional[str] = None) -> str:
        """
        Export scan results to CSV format.

        Args:
            scan_results: Dictionary of open ports and their detailed information
            host: The hostname or IP address scanned
            filename: Optional filename for the export

        Returns:
            str: Path to the exported file
        """
        try:
            # Generate filename if not provided
            if filename is None:
                filename = f"{host}_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

            # Create full path
            filepath = os.path.join(self.export_dir, filename)

            # Prepare data
            data = self._prepare_data(scan_results, host)

            # Write to CSV
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerows(data)

            logger.info(f"Scan results exported to CSV: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            print(f"{Fore.RED}[ERROR] Failed to export to CSV: {e}")
            return ""

    def validate_filename(self, filename: str) -> str:
        """
        Validate and sanitize the export filename.
        
        Args:
            filename: The proposed filename
            
        Returns:
            str: Sanitized filename
            
        Raises:
            ValueError: If filename is invalid
        """
        if not filename:
            raise ValueError("Filename cannot be empty")
        
        # Remove any directory traversal attempts
        filename = os.path.basename(filename)
        
        # Remove any potentially dangerous characters
        filename = re.sub(r'[^\w\-_\.]', '_', filename)
        
        # Ensure the filename has an extension
        if not os.path.splitext(filename)[1]:
            raise ValueError("Filename must have an extension")
        
        return filename

    def ensure_export_directory(self) -> None:
        """
        Ensure the export directory exists and is writable.
        
        Raises:
            OSError: If directory cannot be created or is not writable
        """
        try:
            os.makedirs(self.export_dir, exist_ok=True)
            # Test if directory is writable
            test_file = os.path.join(self.export_dir, '.test')
            try:
                with open(test_file, 'w') as f:
                    f.write('test')
                os.remove(test_file)
            except (IOError, OSError) as e:
                raise OSError(f"Export directory is not writable: {e}")
        except OSError as e:
            raise OSError(f"Cannot create export directory: {e}")

    def export_to_excel(self, scan_results: Dict[int, Dict], host: str, filename: Optional[str] = None) -> str:
        """
        Export scan results to Excel format with enhanced error handling.
        
        Args:
            scan_results: Dictionary of open ports and their detailed information
            host: The hostname or IP address scanned
            filename: Optional filename for the export
            
        Returns:
            str: Path to the exported file
            
        Raises:
            ValueError: If input parameters are invalid
            OSError: If file operations fail
        """
        if not EXCEL_AVAILABLE or openpyxl is None:
            raise ImportError("Excel export not available. Please install openpyxl package.")
        
        if not scan_results:
            raise ValueError("No scan results to export")
        
        # Generate filename if not provided
        if filename is None:
            filename = f"{host}_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        try:
            # Validate and sanitize filename
            filename = self.validate_filename(filename)
            
            # Ensure export directory exists and is writable
            self.ensure_export_directory()
            
            # Create full path
            filepath = os.path.join(self.export_dir, filename)
            
            # Check if file already exists
            if os.path.exists(filepath):
                backup_count = 1
                while os.path.exists(f"{filepath}.{backup_count}"):
                    backup_count += 1
                filepath = f"{filepath}.{backup_count}"
            
            # Create workbook and add data
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Scan Results"
            
            # Add headers
            headers = ["Port", "Status", "Service", "Banner", "SSL Info"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add data
            for row, (port, info) in enumerate(scan_results.items(), 2):
                ws.cell(row=row, column=1, value=port)
                ws.cell(row=row, column=2, value="Open")
                ws.cell(row=row, column=3, value=info.get('service', ''))
                ws.cell(row=row, column=4, value=info.get('banner', ''))
                
                # Format SSL info if available
                ssl_info = info.get('ssl_cert', {})
                if ssl_info:
                    ssl_text = f"Issued to: {ssl_info.get('issued_to', 'Unknown')}\n"
                    ssl_text += f"Issued by: {ssl_info.get('issued_by', 'Unknown')}\n"
                    ssl_text += f"Valid from: {ssl_info.get('valid_from', '')}\n"
                    ssl_text += f"Valid until: {ssl_info.get('valid_until', '')}"
                    ws.cell(row=row, column=5, value=ssl_text)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filepath)
            return filepath
            
        except Exception as e:
            raise OSError(f"Failed to export to Excel: {str(e)}")

    def export_to_pdf(self, scan_results: Dict[int, Dict], host: str, filename: Optional[str] = None) -> str:
        """
        Export scan results to PDF format.

        Args:
            scan_results: Dictionary of open ports and their detailed information
            host: The hostname or IP address scanned
            filename: Optional filename for the export

        Returns:
            str: Path to the exported file
        """
        if not PDF_AVAILABLE or FPDF is None:
            logger.warning("PDF export not available. Install fpdf package.")
            print(f"{Fore.YELLOW}[WARNING] PDF export not available. Install 'fpdf' package.")
            return ""

        try:
            # Ensure the export directory exists
            os.makedirs(self.export_dir, exist_ok=True)
            # Generate filename if not provided
            if filename is None:
                filename = f"{host}_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

            # Create full path
            filepath = os.path.join(self.export_dir, filename)

            # Prepare data
            data = self._prepare_data(scan_results, host)

            # Create PDF object
            pdf = FPDF()
            pdf.add_page()

            # Set font
            pdf.set_font("Arial", size=12)

            # Add title
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(200, 10, f"Port Scan Results for {host}", ln=True, align='C')

            # Add scan timestamp
            pdf.set_font("Arial", size=10)
            pdf.cell(200, 10, f"Scan Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)

            # Add some space
            pdf.ln(10)

            # Select limited columns for PDF display
            # PDF has limited width, so we'll only include key information
            pdf_headers = ["Host", "Port", "Status", "Service", "Version", "Server"]
            pdf_column_indices = [data[0].index(header) for header in pdf_headers if header in data[0]]

            # Set table header
            pdf.set_font("Arial", 'B', 12)
            col_width = min(190 / len(pdf_column_indices), 36)
            row_height = 10

            # Add header row
            for idx in pdf_column_indices:
                pdf.cell(col_width, row_height, data[0][idx], border=1)
            pdf.ln(row_height)

            # Add data rows
            pdf.set_font("Arial", size=8)
            for row in data[1:]:
                for idx in pdf_column_indices:
                    # Truncate and clean text for PDF
                    text = str(row[idx]).replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
                    # Remove non-ASCII characters that might cause PDF issues
                    text = ''.join(char if ord(char) < 128 else '?' for char in text)
                    if len(text) > 40:
                        text = text[:37] + "..."
                    pdf.cell(col_width, row_height, text, border=1)
                pdf.ln(row_height)

            # Save the PDF
            pdf.output(filepath)

            logger.info(f"Scan results exported to PDF: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Error exporting to PDF: {e}")
            print(f"{Fore.RED}[ERROR] Failed to export to PDF: {e}")
            return ""

    def export_to_json(self, scan_results: Dict[int, Dict], host: str, filename: Optional[str] = None) -> str:
        """
        Export scan results to JSON format.

        Args:
            scan_results: Dictionary of open ports and their detailed information
            host: The hostname or IP address scanned
            filename: Optional filename for the export

        Returns:
            str: Path to the exported file
        """
        try:
            # Generate filename if not provided
            if filename is None:
                filename = f"{host}_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

            # Create full path
            filepath = os.path.join(self.export_dir, filename)

            # Create export data structure
            export_data = {
                "scan_info": {
                    "host": host,
                    "scan_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    "total_open_ports": len(scan_results)
                },
                "open_ports": scan_results
            }

            # Write to JSON file
            with open(filepath, 'w', encoding='utf-8') as jsonfile:
                json.dump(export_data, jsonfile, indent=2, default=str, ensure_ascii=False)

            logger.info(f"Scan results exported to JSON: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            print(f"{Fore.RED}[ERROR] Failed to export to JSON: {e}")
            return ""

    def store_export_history(self, 
                           scan_id: str, 
                           target_host: str, 
                           export_format: str, 
                           file_path: str, 
                           user_id: Optional[str] = None,
                           scan_results: Optional[Dict] = None) -> bool:
        """
        Store export information in Supabase.
        
        Args:
            scan_id: Unique ID for the scan
            target_host: Host that was scanned
            export_format: Format of the export (csv, excel, pdf, json)
            file_path: Path to the exported file
            user_id: UUID of the user who performed the export
            scan_results: The scan results data
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Generate a summary from scan results
            summary = "Port scan results"
            port_count = 0
            open_port_count = 0
            
            if scan_results:
                port_count = len(scan_results)
                # Make sure we only count open ports
                open_port_count = sum(1 for p in scan_results if isinstance(scan_results[p], dict) and scan_results[p].get("status") == "open")
                
                # If no status field, assume all are open (legacy format)
                if open_port_count == 0 and port_count > 0:
                    open_port_count = port_count
                
                # Create a short summary of findings
                services = []
                for p in scan_results:
                    port_data = scan_results[p]
                    if isinstance(port_data, dict):
                        service = port_data.get("service", "")
                        if service and service not in services:
                            services.append(service)
                    elif isinstance(port_data, str) and port_data and port_data not in services:
                        services.append(port_data)
                
                top_services = ", ".join(sorted(set([s for s in services if s]))[:5])
                summary = f"Found {open_port_count} open ports out of {port_count} scanned. " + \
                          (f"Top services: {top_services}" if top_services else "")
            
            # Get file size
            file_size = 0
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                
            # Create export record data
            export_data = {
                "scan_id": scan_id,
                "target_host": target_host,
                "export_format": export_format,
                "file_path": file_path,
                "file_size": file_size,
                "scan_date": datetime.now().isoformat(),
                "export_date": datetime.now().isoformat(),
                "port_count": port_count,
                "open_port_count": open_port_count,
                "summary": summary
            }
            
            # Add user_id if provided
            if user_id:
                export_data["user_id"] = user_id
                
            # Get Supabase client
            try:
                supabase = get_supabase()
                logger.info(f"Connected to Supabase, storing export history for scan {scan_id}")
                
                # Store in Supabase
                result = supabase.table("scan_exports").insert(export_data).execute()
                
                if result and hasattr(result, 'data') and len(result.data) > 0:
                    logger.info(f"Successfully stored export history in Supabase: {result.data}")
                    return True
                else:
                    logger.warning(f"Supabase returned empty result: {result}")
                    return False
                
            except Exception as supabase_error:
                logger.error(f"Supabase client error: {str(supabase_error)}")
                print(f"{Fore.RED}[ERROR] Supabase client error: {str(supabase_error)}")
                raise
            
        except Exception as e:
            logger.error(f"Error storing export history: {e}")
            print(f"{Fore.RED}[ERROR] Failed to store export history: {e}")
            return False